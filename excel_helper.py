import openpyxl
from openpyxl.styles import Font, Color
import datetime

import os
from openpyxl.utils import column_index_from_string, get_column_letter
from last_nest import LastNest
from app_data import CONST_LAST_NEST_IDS, CONST_LAST_NEST_INFO

def get_next_column_letter(column_letter, increment=1):    
    column_num = column_index_from_string(column_letter)
    next_column_num = column_num + increment
    next_column_letter = get_column_letter(next_column_num)
    return next_column_letter

class LTdata:
    def __init__(self):
        self.nest_to_row = {}

class LTexcelDef:
    def __init__(self, worksheet):
        #self.nest_id_column = False
        self.worksheet = worksheet
        self.first_empty_column = worksheet.max_column+1
        self.error_msg = []

    def getCell(self, col, row):
        return worksheet[f'{get_column_letter(col)}{row}']

    def append_missing_nest_msg(self, nest_id, line, eggs):
        self.error_msg.append(line)

        #if self.nest_id_column:
        cell = excelLT.getCell(excelLT.first_empty_column + 4, len(self.error_msg) + 1)
        cell.value = nest_id
        cell = excelLT.getCell(excelLT.first_empty_column + 5, len(self.error_msg) + 1)
        cell.value = eggs

        column_offset = 5
        #if self.nest_id_column:
        column_offset = 6
        cell = excelLT.getCell(excelLT.first_empty_column + column_offset, len(self.error_msg) + 1)
        cell.value = line

    def append_error_line_msg(self, line):
        column_offset = 6
        #if self.nest_id_column:
        column_offset = 7

        self.error_msg.append(line)
        cell = excelLT.getCell(excelLT.first_empty_column + column_offset, len(self.error_msg) + 1)
        cell.value = line

    def append_match_nest(self, matched_nest_row, nest_id, line, eggs):
        cell = excelLT.getCell(excelLT.first_empty_column + 1, matched_nest_row)
        cell.value = nest_id

        cell = excelLT.getCell(excelLT.first_empty_column + 2, matched_nest_row)
        if "一雛一蛋" in line or "一蛋一雛" in line or "1蛋1雛" in line or "1雛1蛋" in line:
            eggs = "1蛋1雛"
        elif "兩雛一蛋" in line or "一蛋兩雛" in line or "1蛋2雛" in line or "2雛1蛋" in line:
            eggs = "1蛋2雛"
        elif "一雛兩蛋" in line or "兩蛋一雛" in line or "2蛋1雛" in line or "1雛2蛋" in line:
            eggs = "2蛋1雛"
        if "成功" in str(cell.value) and len(eggs)==0 and not "失敗" in line:
            pass
        elif "失敗" in line:
            cell.value = "失敗"
        elif "成功" in line:
            cell.value = "成功"
        else:
            if "成功" in str(cell.value):
                print("Error ! ")
                exit()
            cell.value = eggs
        egg_cell = cell.value
        #print(f"[Debug] {line} - {cell.value}")
        cell = excelLT.getCell(excelLT.first_empty_column + 3, matched_nest_row)
        if cell.value==None:
            cell.value = ""

        cell.value = str(cell.value) + ("\n" if len(str(cell.value)) > 0 else "") + line
        #print(f"[{matched_nest_row}] {egg_cell}", " -- ", cell.value)

excel_filename = 'data/2023蛋數增長紀錄0702-臨時0709整理.xlsx'
excel_filename = 'data/2023 南澳溪口小燕鷗調查紀錄表 (0716整理).xlsx' # 南澳

line_msg_filename = 'data/LT-line-msg.txt'


dataLT = LTdata()
# Load the workbook and select the first worksheet
workbook = openpyxl.load_workbook(excel_filename)
worksheet = workbook.active

excelLT = LTexcelDef(worksheet)
#excelLT.nest_id_column = True

columns = []
columns.append("巢位")
columns.append("舊巢蛋數")
columns.append("舊巢訊息")
#if excelLT.nest_id_column:
columns.append("其他巢位")
columns.append("蛋數")
columns.append("其他巢位訊息")
columns.append("看不懂的訊息")

for i in range(len(columns)):    
    cell = excelLT.getCell(excelLT.first_empty_column + 1 + i, 1)
    cell.value = columns[i]

with open(line_msg_filename, 'r') as file:
    line_msgs = [_.strip() for _ in file.readlines()]

last_nest = LastNest()
#last_nest.setData(CONST_LAST_NEST_INFO, CONST_LAST_NEST_IDS)
#last_nest.parseInfo()

COLUMN_NEST_ID = -1
for i in range(worksheet.max_column):    
    cell = excelLT.getCell(i + 1, 1)
    if "編號" in ("" + str(cell.value)):
        COLUMN_NEST_ID = i+1
if COLUMN_NEST_ID <= 0:
    print("error 找不到名稱包含 編號 的欄位")
    exit()
COLUMN_NEST_ID = get_column_letter(COLUMN_NEST_ID)
print("編號欄位是", COLUMN_NEST_ID)

for cell in worksheet[COLUMN_NEST_ID + '1':COLUMN_NEST_ID + '2000']:
    if cell[0].value is not None:
        if cell[0].value in dataLT.nest_to_row:
            print(f"[error] 巢號重覆 {cell[0].value} 第 {cell[0].row} 列 ")
        dataLT.nest_to_row[cell[0].value] = str(cell[0].row)
        #if not cell[0].value in last_nest.last_ids and not cell[0].value == "編號":
        #    print(f"[warning] app_data.py 巢號缺少 {cell[0].value} ")

for line in line_msgs:
    line_arr = line.split(' ')
    if line.split(' ')[-1] == "圖片":
        pass
    elif len(line_arr) > 1 and line_arr[1] == "小燕鷗調查小幫手":
        pass
    else:
        if "小燕鷗調查小幫手" in line and "重新定" in line:
            print("debug line ; ", line)
            print("debug, ", line_arr,  len(line_arr) > 1 and line_arr[1] == "小燕鷗調查小幫手")
            print("found 小燕鷗調查小幫手", len(line_arr), line_arr[1])
            exit()

        nest_id = last_nest.get_line_nest_id(line)

        matched_nest_row = 0
        if nest_id in dataLT.nest_to_row:
            matched_nest_row = dataLT.nest_to_row[nest_id]
        
        if len(nest_id) == 0:
            #看不懂的訊息 沒有已知的巢位編號
            if len(line_arr) > 1 and line_arr[1] == "小燕鷗調查小幫手":
                pass
            elif len(line_arr) > 1 and "已收回訊息" in line_arr[1]:
                pass
            else:
                excelLT.append_error_line_msg(line)

        elif matched_nest_row == 0:
            #有巢位訊息但是excel裡面沒有該巢            
            eggs = last_nest.get_line_eggs(nest_id, line, to_int=True)            
            excelLT.append_missing_nest_msg(nest_id, line, eggs)
            

        else:
            #有巢位訊息可以對應到excel的資料列
            eggs = last_nest.get_line_eggs(nest_id, line, to_int=True)
            excelLT.append_match_nest(matched_nest_row, nest_id, line, eggs)

# Save the changes
date_str = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
workbook.save(f'data/LT-output-last.xlsx')

#bash delete all output 
#find ./data/ -type f -name "*LT-output*.xlsx*" | xargs -d"\n" rm -f